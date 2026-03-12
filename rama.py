"""
Rwanda Health Voucher — Multi-Clinic Fraud Detection Dashboard
══════════════════════════════════════════════════════════════

INSTALL (once):
    pip install streamlit pandas openpyxl plotly

RUN:
    streamlit run ndengera_streamlit.py

Supports any number of clinics uploaded simultaneously.
Clinic name, district & TIN are auto-detected from each file header.
"""

import re, io
from datetime import datetime, date
from difflib import SequenceMatcher

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# ─────────────────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Rwanda Voucher Fraud Dashboard",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
#  CSS — FULL DARK MODE THEME
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
/* ── Global background & text ───────────────────────────────────────────── */
html, body, [data-testid="stAppViewContainer"],
[data-testid="stMain"], .main, .block-container {
    background-color: #0D1117 !important;
    color: #E6EDF3 !important;
}
[data-testid="stAppViewContainer"] > .main { background-color: #0D1117 !important; }
.block-container { padding-top: 1rem !important; }

/* ── Sidebar ─────────────────────────────────────────────────────────────── */
[data-testid="stSidebar"] {
    background-color: #161B22 !important;
    border-right: 1px solid #30363D !important;
}
[data-testid="stSidebar"] * { color: #C9D1D9 !important; }
[data-testid="stSidebar"] .stMarkdown h3 { color: #58A6FF !important; }

/* ── Tabs ────────────────────────────────────────────────────────────────── */
[data-testid="stTabs"] { background: transparent; }
button[data-baseweb="tab"] {
    background: #161B22 !important;
    color: #8B949E !important;
    border-bottom: 2px solid transparent !important;
    font-weight: 600 !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color: #58A6FF !important;
    border-bottom: 2px solid #58A6FF !important;
    background: #0D1117 !important;
}
[data-testid="stTabPanel"] { background: #0D1117 !important; }

/* ── Metric cards ────────────────────────────────────────────────────────── */
[data-testid="metric-container"] {
    background: #161B22 !important;
    border: 1px solid #30363D !important;
    border-radius: 10px !important;
    padding: .75rem .9rem !important;
}
[data-testid="metric-container"] label,
[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    color: #8B949E !important; font-size: .78rem !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    color: #E6EDF3 !important; font-size: 1.35rem !important; font-weight: 700 !important;
}
[data-testid="metric-container"] [data-testid="stMetricDelta"] { font-size: .78rem !important; }

/* ── Inputs: selects, text inputs ───────────────────────────────────────── */
[data-testid="stSelectbox"] > div > div,
[data-testid="stMultiSelect"] > div > div,
[data-baseweb="select"] > div,
[data-baseweb="input"] > div {
    background-color: #21262D !important;
    border: 1px solid #30363D !important;
    color: #E6EDF3 !important;
    border-radius: 7px !important;
}
input, textarea { background: #21262D !important; color: #E6EDF3 !important; }
[data-baseweb="popover"], [data-baseweb="menu"] {
    background: #21262D !important; border: 1px solid #30363D !important;
}
li[role="option"] { background: #21262D !important; color: #C9D1D9 !important; }
li[role="option"]:hover { background: #30363D !important; }

/* ── Labels above filters ───────────────────────────────────────────────── */
[data-testid="stSelectbox"] label,
[data-testid="stMultiSelect"] label,
[data-testid="stTextInput"] label,
[data-testid="stFileUploader"] label { color: #8B949E !important; font-size:.8rem !important; }

/* ── Buttons ─────────────────────────────────────────────────────────────── */
[data-testid="baseButton-primary"] {
    background: #1F6FEB !important; color: #fff !important;
    border: none !important; border-radius: 7px !important;
}
[data-testid="baseButton-secondary"],
[data-testid="baseButton-secondaryFormSubmit"] {
    background: #21262D !important; color: #C9D1D9 !important;
    border: 1px solid #30363D !important; border-radius: 7px !important;
}
[data-testid="stDownloadButton"] button {
    background: #1F6FEB !important; color: #fff !important;
    border: none !important; font-weight: 600 !important;
}

/* ── Expander ────────────────────────────────────────────────────────────── */
[data-testid="stExpander"] {
    background: #161B22 !important; border: 1px solid #30363D !important;
    border-radius: 8px !important;
}
[data-testid="stExpander"] summary { color: #C9D1D9 !important; }

/* ── Progress / spinner / info / success ────────────────────────────────── */
[data-testid="stInfo"]    { background: #1C2A3A !important; color: #58A6FF !important; border-color: #1F6FEB !important; }
[data-testid="stSuccess"] { background: #1A2E22 !important; color: #3FB950 !important; border-color: #238636 !important; }
[data-testid="stWarning"] { background: #2E2218 !important; color: #D29922 !important; border-color: #9E6A03 !important; }

/* ── Dividers ────────────────────────────────────────────────────────────── */
hr { border-color: #30363D !important; }

/* ── Caption / small text ───────────────────────────────────────────────── */
[data-testid="stCaptionContainer"], small, .stCaption {
    color: #8B949E !important;
}
p, li { color: #C9D1D9 !important; }
h1,h2,h3,h4 { color: #E6EDF3 !important; }

/* ── File uploader ───────────────────────────────────────────────────────── */
[data-testid="stFileUploader"] > div {
    background: #161B22 !important; border: 1px dashed #30363D !important;
    border-radius: 8px !important;
}

/* ── Scrollbar ───────────────────────────────────────────────────────────── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0D1117; }
::-webkit-scrollbar-thumb { background: #30363D; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #58A6FF; }

/* ══ CUSTOM COMPONENTS ═══════════════════════════════════════════════════════*/

/* Header banner */
.main-header {
    background: linear-gradient(135deg,#0D2137 0%,#1A3C5E 60%,#1F6FEB 100%);
    color:#fff; border-radius:12px; padding:1.1rem 1.6rem;
    margin-bottom:1rem; display:flex; align-items:center; gap:1rem;
    border: 1px solid #1F6FEB44;
    box-shadow: 0 4px 24px rgba(31,111,235,.2);
}
.main-header h1 { margin:0; font-size:1.3rem; font-weight:800; letter-spacing:.4px; }
.main-header p  { margin:0; font-size:.8rem; opacity:.75; }

/* Section titles */
.section-title {
    font-size:.92rem; font-weight:700; color:#58A6FF;
    border-left:3px solid #1F6FEB; padding-left:.6rem; margin-bottom:.7rem;
    letter-spacing:.2px;
}

/* Clinic badges */
.clinic-badge {
    display:inline-block; background:#1C2A3A; border:1px solid #1F6FEB66;
    border-radius:20px; padding:.2rem .8rem; font-size:.78rem;
    font-weight:600; color:#58A6FF; margin:.15rem;
}

/* Duplicate type chips */
.chip { display:inline-block; border-radius:6px;
        padding:.3rem .85rem; font-size:.78rem; font-weight:700; margin:.15rem;
        letter-spacing:.2px; }
.chip-exact   { background:#C0392B; color:#fff; }
.chip-voucher { background:#1A6EA8; color:#fff; }
.chip-clinic  { background:#7D3C98; color:#fff; }
.chip-cross   { background:#CA6F1E; color:#fff; }
.chip-month   { background:#1E8449; color:#fff; }

/* Guide cards */
.guide-card {
    background: #161B22; border: 1px solid #30363D; border-radius: 10px;
    padding: 1rem 1.2rem; margin-bottom: .75rem;
}
.guide-card-title {
    font-size: .88rem; font-weight: 700; color: #58A6FF; margin-bottom: .4rem;
}
.guide-card p { color: #C9D1D9 !important; font-size: .84rem; margin: .2rem 0; }
.guide-tip {
    background: #1C2A3A; border-left: 3px solid #1F6FEB;
    border-radius: 0 6px 6px 0; padding: .5rem .8rem;
    font-size: .82rem; color: #79C0FF !important; margin-top: .5rem;
}

/* Upload hint */
.upload-hint {
    background:#0D2137; border:2px dashed #1F6FEB66;
    border-radius:12px; padding:2.5rem 2rem; text-align:center;
    color:#58A6FF; font-size:.9rem;
}
.upload-hint h3 { color:#79C0FF !important; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
SERVICE_COLS = [
    "Consultation", "Laboratory", "Imaging",
    "Hospitalization", "Procedures", "Consumables", "Medicines",
]

DUP_COLORS = {
    "Exact Duplicate":        "#C0392B",   # deep red   — most critical
    "Duplicate Voucher":      "#1A6EA8",   # deep blue
    "Cross-Clinic Duplicate": "#7D3C98",   # deep purple — high risk
    "Cross-Month Duplicate":  "#CA6F1E",   # deep orange
    "Same ID Same Month":     "#1E8449",   # deep green
}

# Text color to pair with each background (always white for dark fills)
DUP_TEXT = {k: "#FFFFFF" for k in DUP_COLORS}

DUP_ORDER = list(DUP_COLORS.keys())

CHART_COLORS = [
    "#2471A3","#1ABC9C","#F39C12","#E74C3C",
    "#9B59B6","#27AE60","#E67E22","#17A589",
    "#2E86C1","#D35400",
]

# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS — PARSING
# ─────────────────────────────────────────────────────────────────────────────
def eval_cost(val) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if re.match(r"=\s*SUM", s, re.IGNORECASE):
        inner = re.search(r"\((.+)\)", s)
        if inner:
            return sum(float(n) for n in re.findall(r"[\d.]+", inner.group(1)))
    try:
        return float(s)
    except Exception:
        return 0.0


def parse_date(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    s = str(val).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except Exception:
            pass
    return s


def normalize_aff(name: str) -> str:
    s = str(name).upper()
    s = re.sub(r"[/\-_,.]", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def fuzzy_canonical(name: str, groups: list, threshold: float = 0.78) -> str:
    norm = normalize_aff(name)
    if not norm:
        return "UNKNOWN"
    best, best_score = None, 0.0
    for g in groups:
        sc = SequenceMatcher(None, norm, normalize_aff(g)).ratio()
        if sc > best_score:
            best_score, best = sc, g
    return best if best is not None and best_score >= threshold else norm


# ─────────────────────────────────────────────────────────────────────────────
#  METADATA EXTRACTION  (clinic name, district, TIN, invoice month)
# ─────────────────────────────────────────────────────────────────────────────
def extract_metadata(rows: list) -> dict:
    meta = {"clinic": "", "district": "", "tin": "", "month": ""}

    def flatten(row) -> str:
        return " ".join(str(c).strip() for c in row if c).upper()

    # Scan full row list — some files put values in adjacent cells
    for i, row in enumerate(rows[:10]):
        line = flatten(row)
        if not line:
            continue

        # "HEALTH FACILITY" keyword — value may be in same cell or next cell
        if "HEALTH FACILITY" in line:
            # Try to parse value after the colon in the combined line
            m = re.search(r"HEALTH\s+FACILITY[:\s]+(.+)", line)
            if m and m.group(1).strip():
                meta["clinic"] = m.group(1).strip().title()
            else:
                # Value might be in the cell right after (column B onward)
                for cell in row[1:]:
                    if cell and str(cell).strip():
                        meta["clinic"] = str(cell).strip().title()
                        break

        if "DISTRICT" in line:
            m = re.search(r"DISTRICT[:\s]+(.+)", line)
            if m and m.group(1).strip():
                meta["district"] = m.group(1).strip().title()
            else:
                for cell in row[1:]:
                    if cell and str(cell).strip():
                        meta["district"] = str(cell).strip().title()
                        break

        m = re.search(r"TIN\s*(?:NO)?[:\s]+([\d]+)", line)
        if m:
            meta["tin"] = m.group(1)

        m = re.search(r"INVOICE\s+OF\s+(.+)", line)
        if m:
            meta["month"] = m.group(1).strip().title()

    return meta


# Abbreviation map for filename-based clinic name fallback
_ABBREV = {
    "PC":  "Polyclinic",
    "HC":  "Health Centre",
    "DH":  "District Hospital",
    "CSB": "Centre de Santé de Base",
    "HP":  "Hôpital",
}

def clinic_from_filename(filename: str) -> str:
    """
    Derive a clean clinic name from a filename like:
      NDENGERA-PC-2025-SEPTEMBER.xlsx  →  Ndengera Polyclinic
      BIRUNGA-HC-2025-AUGUST.xlsx      →  Birunga Health Centre
    """
    stem = re.sub(r"\.[a-zA-Z]+$", "", filename)   # strip extension
    stem = re.sub(r"^\d+_", "", stem)               # strip numeric prefix
    # Remove year and month tokens
    stem = re.sub(r"\b(20\d{2})\b", "", stem)
    months = ("JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE",
              "JULY","AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER")
    for mo in months:
        stem = re.sub(mo, "", stem, flags=re.IGNORECASE)
    parts = [p.strip() for p in re.split(r"[-_\s]+", stem) if p.strip()]
    # Expand known abbreviations
    expanded = []
    for p in parts:
        expanded.append(_ABBREV.get(p.upper(), p.title()))
    return " ".join(expanded).strip() or filename


# ─────────────────────────────────────────────────────────────────────────────
#  COLUMN LAYOUT DETECTION  — finds where "DATE" header lives
# ─────────────────────────────────────────────────────────────────────────────
def find_data_start(rows: list):
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if str(cell).strip().upper() == "DATE":
                return i, j
    return 9, 1  # fallback


# ─────────────────────────────────────────────────────────────────────────────
#  FILE PARSING
# ─────────────────────────────────────────────────────────────────────────────
def parse_file(file_bytes: bytes, filename: str) -> pd.DataFrame:
    wb   = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    meta    = extract_metadata(rows)
    hdr_idx, date_col = find_data_start(rows)

    # Use filename as fallback clinic name if not found in header
    clinic = meta["clinic"] or clinic_from_filename(filename)

    C       = date_col
    records = []

    for row in rows[hdr_idx + 1:]:
        aff_raw = row[C + 6] if len(row) > C + 6 else None
        affil   = str(aff_raw).strip() if aff_raw else ""
        if not affil or affil.upper().startswith("TOTAL") or affil.startswith("="):
            if affil.upper().startswith("TOTAL"):
                break
            continue

        date_val = row[C] if len(row) > C else None
        if date_val is None or str(date_val).strip() in ("", "None", "DATE"):
            continue

        def col(offset):
            idx = C + offset
            return row[idx] if len(row) > idx else None

        rec = {
            "Date":             parse_date(col(0)),
            "Voucher_ID":       str(col(1) or "").strip(),
            "Affiliation_No":   str(col(2) or "").strip(),
            "Age":              str(col(3) or "").strip(),
            "Beneficiary_Name": str(col(4) or "").strip().upper(),
            "Affiliate_Name":   str(col(5) or "").strip().upper(),
            "Affiliation":      affil,
            "Consultation":     eval_cost(col(7)),
            "Laboratory":       eval_cost(col(8)),
            "Imaging":          eval_cost(col(9)),
            "Hospitalization":  eval_cost(col(10)),
            "Procedures":       eval_cost(col(11)),
            "Consumables":      eval_cost(col(12)),
            "Medicines":        eval_cost(col(13)),
            "Clinic":           clinic,
            "District":         meta["district"],
            "TIN":              meta["tin"],
            "Month_Label":      meta["month"] or filename,
            "Source_File":      filename,
        }
        rec["Total_100"] = sum(rec[s] for s in SERVICE_COLS)
        rec["Total_85"]  = round(rec["Total_100"] * 0.85, 0)
        records.append(rec)

    return pd.DataFrame(records)


# ─────────────────────────────────────────────────────────────────────────────
#  AFFILIATION NORMALISATION
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def build_canonical_affiliations(df: pd.DataFrame) -> pd.DataFrame:
    groups: list    = []
    canon_map: dict = {}
    for name in df["Affiliation"].unique():
        canon = fuzzy_canonical(name, groups)
        if canon not in groups:
            groups.append(canon)
        canon_map[name] = canon
    df = df.copy()
    df["Affiliation_Group"] = df["Affiliation"].map(canon_map)
    return df


# ─────────────────────────────────────────────────────────────────────────────
#  DUPLICATE DETECTION
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def detect_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["Dup_Type"] = ""

    # A — Exact: same ID + name + date (any clinic)
    mask_a = df.duplicated(
        subset=["Affiliation_No", "Beneficiary_Name", "Date"], keep=False
    )
    df.loc[mask_a, "Dup_Type"] = "Exact Duplicate"

    # B — Duplicate voucher number
    mask_b = df.duplicated(subset=["Voucher_ID"], keep=False) & (df["Dup_Type"] == "")
    df.loc[mask_b, "Dup_Type"] = "Duplicate Voucher"

    # C — Cross-Clinic: same patient at multiple clinics
    n_clinics = df.groupby(
        ["Affiliation_No", "Beneficiary_Name"]
    )["Clinic"].transform("nunique")
    mask_c = (n_clinics > 1) & (df["Dup_Type"] == "")
    df.loc[mask_c, "Dup_Type"] = "Cross-Clinic Duplicate"

    # D — Cross-Month: same patient in multiple months at same clinic
    key       = df["Affiliation_No"] + "||" + df["Beneficiary_Name"] + "||" + df["Clinic"]
    month_cnt = df.groupby(key.rename("key"))["Month_Label"].transform("nunique")
    mask_d    = (month_cnt > 1) & (df["Dup_Type"] == "")
    df.loc[mask_d, "Dup_Type"] = "Cross-Month Duplicate"

    # E — Same ID same month at same clinic
    per_month = df.groupby(
        ["Affiliation_No", "Month_Label", "Clinic"]
    )["Beneficiary_Name"].transform("count")
    mask_e = (per_month > 1) & (df["Dup_Type"] == "")
    df.loc[mask_e, "Dup_Type"] = "Same ID Same Month"

    return df


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def build_export(df: pd.DataFrame) -> bytes:
    buf    = io.BytesIO()
    wb_out = openpyxl.Workbook()

    h_fill = PatternFill("solid", fgColor="1A3C5E")
    h_font = Font(color="FFFFFF", bold=True)
    # Deep fills matching dashboard colors — readable on screen and in print
    fills = {k: PatternFill("solid", fgColor=v.lstrip("#"))
             for k, v in DUP_COLORS.items()}
    white_font = Font(color="FFFFFF", bold=False)

    def write_df(ws, data_df, color_dup=False):
        cols    = list(data_df.columns)
        dup_idx = cols.index("Dup_Type") if "Dup_Type" in cols else None
        for c, col_name in enumerate(cols, 1):
            cell           = ws.cell(1, c, col_name.replace("_", " "))
            cell.fill      = h_fill
            cell.font      = h_font
            cell.alignment = Alignment(horizontal="center")
        for r, row_vals in enumerate(data_df.itertuples(index=False), 2):
            for c, val in enumerate(row_vals, 1):
                ws.cell(r, c, "" if (isinstance(val, float) and pd.isna(val)) else val)
            if color_dup and dup_idx is not None:
                dtype = row_vals[dup_idx]
                if dtype in fills:
                    for c2 in range(1, len(cols) + 1):
                        ws.cell(r, c2).fill = fills[dtype]
                        ws.cell(r, c2).font = white_font
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(w + 2, 50)

    # Sheet 1 — Full data
    ws1 = wb_out.active
    ws1.title = "Full Data"
    write_df(ws1, df, color_dup=True)

    # Sheet 2 — Duplicates
    ws2      = wb_out.create_sheet("Duplicates")
    dup_cols = [
        "Date", "Voucher_ID", "Affiliation_No", "Beneficiary_Name",
        "Clinic", "District", "Affiliation_Group",
        "Total_100", "Total_85", "Dup_Type", "Month_Label", "Source_File",
    ]
    dup_df = (df[df["Dup_Type"] != ""][dup_cols].copy()
              if (df["Dup_Type"] != "").any()
              else pd.DataFrame(columns=dup_cols))
    write_df(ws2, dup_df, color_dup=True)

    # Sheet 3 — By Affiliation per Clinic
    ws3    = wb_out.create_sheet("By Affiliation")
    aff_df = (df.groupby(["Clinic", "Affiliation_Group"]).agg(
                  Total_Claims    = ("Voucher_ID",    "count"),
                  Total_100       = ("Total_100",     "sum"),
                  Total_85        = ("Total_85",      "sum"),
                  Flagged_Entries = ("Dup_Type",      lambda x: (x != "").sum()),
                  Months_Active   = ("Month_Label",   "nunique"),
              ).reset_index()
              .sort_values(["Clinic", "Total_100"], ascending=[True, False]))
    write_df(ws3, aff_df)

    # Sheet 4 — By Month per Clinic
    ws4      = wb_out.create_sheet("By Month")
    month_df = (df.groupby(["Clinic", "Month_Label"]).agg(
                    Total_Claims        = ("Voucher_ID",        "count"),
                    Total_100           = ("Total_100",         "sum"),
                    Total_85            = ("Total_85",          "sum"),
                    Flagged_Entries     = ("Dup_Type",          lambda x: (x != "").sum()),
                    Unique_Affiliations = ("Affiliation_Group", "nunique"),
                ).reset_index())
    write_df(ws4, month_df)

    # Sheet 5 — Cross-Clinic flags
    ws5   = wb_out.create_sheet("Cross-Clinic Flags")
    cc_df = (df[df["Dup_Type"] == "Cross-Clinic Duplicate"][[
                "Date", "Voucher_ID", "Affiliation_No", "Beneficiary_Name",
                "Clinic", "District", "Affiliation_Group",
                "Total_100", "Total_85", "Month_Label", "Source_File",
            ]].copy()
             .sort_values(["Affiliation_No", "Beneficiary_Name", "Date"]))
    write_df(ws5, cc_df, color_dup=False)
    cc_fill = PatternFill("solid", fgColor="7D3C98")
    for row in ws5.iter_rows(min_row=2, max_row=ws5.max_row):
        for cell in row:
            cell.fill = cc_fill
            cell.font = white_font

    wb_out.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  DISPLAY HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def fmt(n) -> str:
    return f"RWF {int(n):,}"


def highlight_dups(row):
    bg    = DUP_COLORS.get(row.get("Dup_Type", ""), "")
    if bg:
        style = f"background-color:{bg}; color:#ffffff; font-weight:500;"
        return [style for _ in row]
    return ["" for _ in row]


# ─────────────────────────────────────────────────────────────────────────────
#  GUIDE RENDERER
# ─────────────────────────────────────────────────────────────────────────────
def _render_guide():
    st.markdown('<div class="section-title">📖 User Guide — Filters, Sorts & Duplicate Types</div>',
                unsafe_allow_html=True)

    # ── Section 1: Duplicate Types ───────────────────────────────────────────
    st.markdown("""
    <div class="guide-card">
      <div class="guide-card-title">🔴 DUPLICATE TYPE DEFINITIONS</div>
      <p>The system automatically scans every uploaded record and assigns one of five fraud/anomaly flags.
         A record gets at most <strong>one flag</strong> — the most severe match wins.</p>
    </div>
    """, unsafe_allow_html=True)

    dup_guide = [
        ("#C0392B", "⚠ Exact Duplicate",
         "Same <strong>Affiliation Number + Beneficiary Name + Date</strong> appears more than once.",
         "This is the clearest sign of double-billing. The same patient visit has been submitted twice "
         "(or more). Check whether the voucher IDs are also the same — if so, it is likely a data entry "
         "error; if the voucher IDs differ, it suggests intentional re-submission for payment.",
         "Investigate immediately. One of the entries must be rejected."),
        ("#1A6EA8", "⚠ Duplicate Voucher",
         "The same <strong>Voucher ID</strong> appears more than once across all records.",
         "Each voucher should be a unique reference number for one patient visit at one facility. "
         "Seeing the same voucher number twice means either the voucher was re-used fraudulently, "
         "scanned/entered twice by accident, or the numbering system at the clinic has an error.",
         "Cross-check the physical voucher with the facility register. Only one claim per voucher ID is valid."),
        ("#7D3C98", "⚠ Cross-Clinic Duplicate",
         "Same <strong>Affiliation Number + Beneficiary Name</strong> appears at <strong>two or more different clinics</strong>.",
         "Under normal circumstances, a patient can visit multiple clinics — but the combination of "
         "the same national ID number and the same full name at multiple facilities within the same "
         "reporting period is a strong fraud indicator. The patient may not actually exist, or one "
         "clinic may be submitting claims for patients seen elsewhere.",
         "Compare visit dates. If dates overlap or are suspiciously close, escalate for investigation."),
        ("#CA6F1E", "ℹ Cross-Month Duplicate",
         "Same <strong>Affiliation Number + Beneficiary Name</strong> appears in <strong>more than one invoice month</strong> at the same clinic.",
         "A patient returning to the same clinic across different months is normal. This flag is "
         "informational — it highlights repeat visitors who may deserve closer scrutiny if combined "
         "with high costs, or if they appear every month with identical service combinations.",
         "Review service patterns. Repeated hospitalisation or imaging every month may warrant follow-up."),
        ("#1E8449", "ℹ Same ID Same Month",
         "Same <strong>Affiliation Number</strong> appears more than once within the <strong>same month at the same clinic</strong>, but with different names.",
         "One affiliation number should belong to one person. If multiple names share the same ID in "
         "the same month, this could mean: (1) a family member sharing a card, (2) a data entry error "
         "in the affiliation number, or (3) a fraudulent claim using someone else's ID.",
         "Check whether the names are similar (possible spelling variation) or completely different (possible ID misuse)."),
    ]

    for color, title, trigger, explanation, action in dup_guide:
        st.markdown(f"""
        <div class="guide-card" style="border-left:4px solid {color}">
          <div class="guide-card-title" style="color:{color}">{title}</div>
          <p><strong>Triggered when:</strong> {trigger}</p>
          <p><strong>What it means:</strong> {explanation}</p>
          <div class="guide-tip">💡 <strong>Recommended action:</strong> {action}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Section 2: Sidebar Filters ───────────────────────────────────────────
    st.markdown("""
    <div class="guide-card">
      <div class="guide-card-title">🔍 GLOBAL SIDEBAR FILTER</div>
    </div>
    """, unsafe_allow_html=True)

    sidebar_filters = [
        ("🏥 Filter by Clinic (sidebar multiselect)",
         "Restricts ALL tabs at once to only show data from the selected clinics. "
         "Deselecting a clinic removes it from every chart, table, and export in the current session. "
         "Use this when you want to compare two specific clinics without the noise of others, "
         "or to focus a full audit on one facility.",
         "Selecting zero clinics shows all data (safety fallback)."),
    ]
    for title, explanation, tip in sidebar_filters:
        st.markdown(f"""
        <div class="guide-card">
          <div class="guide-card-title">{title}</div>
          <p>{explanation}</p>
          <div class="guide-tip">💡 {tip}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Section 3: Duplicates Tab Filters ────────────────────────────────────
    st.markdown("""
    <div class="guide-card">
      <div class="guide-card-title">🚨 DUPLICATES TAB — FILTER GUIDE</div>
      <p>This tab shows only flagged records. All four filters work together (AND logic) —
         applying multiple filters narrows the list further.</p>
    </div>
    """, unsafe_allow_html=True)

    dup_filters = [
        ("Filter: Type",
         "Selects which category of duplicate to display. Choose <em>All</em> to see every flagged record, "
         "or pick a specific type (e.g. <em>Exact Duplicate</em>) to focus on the most severe cases first. "
         "The count shown updates to reflect your selection.",
         "Start your review with Exact Duplicate and Cross-Clinic — these are highest priority."),
        ("Filter: Month",
         "Limits the duplicate list to records from a single invoice month. Useful when you are "
         "auditing one specific month's submission before processing payment.",
         "Compare the same month across clinics to spot if a voucher number was shared."),
        ("Filter: Clinic",
         "Narrows the duplicate list to a single facility. Use this when a specific clinic has "
         "raised concerns or when you want to prepare a per-clinic audit report.",
         "Combine with the Type filter: set Type = Cross-Clinic and Clinic = a specific facility "
         "to see all cross-clinic fraud signals involving that facility."),
        ("Filter: Name Search",
         "Free-text search on the Beneficiary Name column. Type a partial name (case-insensitive) "
         "and only rows containing that string are shown. Useful for following up on a specific "
         "patient flagged during a manual review.",
         "You can search with partial names — e.g. typing 'NIYA' will match NIYOMIZERO, NIYIGENA, etc."),
    ]
    for title, explanation, tip in dup_filters:
        st.markdown(f"""
        <div class="guide-card">
          <div class="guide-card-title">{title}</div>
          <p>{explanation}</p>
          <div class="guide-tip">💡 {tip}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Section 4: Data Explorer Filters ─────────────────────────────────────
    st.markdown("""
    <div class="guide-card">
      <div class="guide-card-title">🔎 DATA EXPLORER TAB — FILTER GUIDE</div>
      <p>Explore the full dataset with up to six simultaneous filters. All filters combine with
         AND logic. The record count updates in real time.</p>
    </div>
    """, unsafe_allow_html=True)

    expl_filters = [
        ("Filter: Clinic",
         "Restricts the explorer to one facility's records only. Combines with the global sidebar "
         "clinic filter — the sidebar narrows the pool, then this filter further restricts within it.",
         "Use with the Service filter to see what a specific clinic claims most for."),
        ("Filter: Month",
         "Shows only records from the chosen invoice month. Allows you to inspect the raw claims "
         "for one billing period before approving payment.",
         "Filter by Month + Flag Status = Flagged Only for a fast pre-payment fraud check."),
        ("Filter: Service Used",
         "Shows only records where the selected service has a non-zero cost. The seven services are: "
         "<strong>Consultation</strong> (clinical review fee), <strong>Laboratory</strong> (blood tests, cultures), "
         "<strong>Imaging</strong> (X-ray, ultrasound, CT), <strong>Hospitalization</strong> (inpatient stays), "
         "<strong>Procedures</strong> (surgical or clinical procedures and materials), "
         "<strong>Consumables</strong> (medical consumables billed separately), "
         "<strong>Medicines</strong> (drugs dispensed). "
         "Filtering to a single service lets you spot outliers — e.g., unusually high imaging costs.",
         "Combine Service = Hospitalization with Flag Status = Flagged Only to find hospitalisation "
         "claims that are also duplicates."),
        ("Filter: Affiliation Group",
         "Filters by the employer/insurer affiliation. The system auto-groups similar spellings "
         "(e.g. 'MINEDUC/RUBAVU' and 'MINEDUC RUBAVU' are merged). Use this to audit all claims "
         "submitted under one employer, which is useful when an employer reports a discrepancy.",
         "Affiliation groups are fuzzy-matched — if you expect more or fewer records than shown, "
         "check the raw Affiliation column for spelling variants."),
        ("Filter: Flag Status",
         "<strong>All</strong> — shows every record including clean ones. "
         "<strong>Flagged Only</strong> — shows only records with any duplicate flag. "
         "<strong>Clean Only</strong> — shows only records with no flags (useful to verify clean submissions). "
         "You can also select a specific duplicate type here as a shortcut.",
         "Use Clean Only before export to quickly verify the proportion of clean claims per clinic."),
        ("Filter: Name Search",
         "Searches the Beneficiary Name field in real time. Partial matches work — type any part "
         "of the name. The search is case-insensitive.",
         "Use to look up a specific individual flagged in the Duplicates tab."),
    ]
    for title, explanation, tip in expl_filters:
        st.markdown(f"""
        <div class="guide-card">
          <div class="guide-card-title">{title}</div>
          <p>{explanation}</p>
          <div class="guide-tip">💡 {tip}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Section 5: Table Sorting ──────────────────────────────────────────────
    st.markdown("""
    <div class="guide-card">
      <div class="guide-card-title">↕ TABLE SORTING</div>
      <p>Every table in the Duplicates and Data Explorer tabs is <strong>sortable by clicking any column header</strong>.
         Click once to sort ascending (A→Z / low→high), click again to sort descending (Z→A / high→low),
         click a third time to reset to the default order.</p>
      <p><strong>Most useful column sorts:</strong></p>
    </div>
    """, unsafe_allow_html=True)

    sort_guide = [
        ("Sort by Total(100%) — Descending",
         "Brings the highest-value claims to the top. Useful for quickly identifying the most expensive "
         "flagged records — large amounts combined with a fraud flag should be investigated first."),
        ("Sort by Date — Ascending",
         "Shows the chronological order of claims. Useful for spotting if the same patient was seen on "
         "the same day at multiple clinics (Cross-Clinic fraud within a single day)."),
        ("Sort by Affiliation_No — Ascending",
         "Groups all records by patient ID, making it easy to see how many times the same ID number "
         "appears across different rows, names, or dates."),
        ("Sort by Beneficiary_Name — Ascending",
         "Alphabetical order. Useful for manually spotting name variations of the same person "
         "(e.g. 'JEAN PIERRE NKURUNZIZA' vs 'J P NKURUNZIZA') that the system may not have auto-matched."),
        ("Sort by Dup_Type — Ascending",
         "Groups all rows by their duplicate category so you can review all Exact Duplicates "
         "together, then all Cross-Clinic flags, etc."),
        ("Sort by Clinic — Ascending",
         "In multi-clinic views, groups all records from the same facility together so you can "
         "compare patterns side by side."),
    ]
    for title, explanation in sort_guide:
        st.markdown(f"""
        <div class="guide-card">
          <div class="guide-card-title">↕ {title}</div>
          <p>{explanation}</p>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("---")

    # ── Section 6: Amount columns ─────────────────────────────────────────────
    st.markdown("""
    <div class="guide-card">
      <div class="guide-card-title">💰 AMOUNT COLUMNS EXPLAINED</div>
      <p><strong>Total(100%)</strong> — The full claimed amount before insurance deduction.
         This is the gross cost of all services in one visit.</p>
      <p><strong>Total(85%)</strong> — The amount payable by the insurer (85% of the 100% total).
         The patient is responsible for the remaining 15% co-payment.
         This column is the actual financial exposure for the insurance fund.</p>
      <div class="guide-tip">💡 When assessing fraud impact, always refer to Total(85%) as that is
         the money the insurer would actually pay. For a duplicated claim,
         the potential over-payment is the Total(85%) of the fraudulent entry.</div>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
#  SESSION STATE
# ─────────────────────────────────────────────────────────────────────────────
if "df" not in st.session_state:
    st.session_state.df           = pd.DataFrame()
    st.session_state.files_loaded = []

# ─────────────────────────────────────────────────────────────────────────────
#  HEADER
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
  <span style="font-size:2rem">🏥</span>
  <div>
    <h1>Rwanda Health Voucher — Fraud Detection Dashboard</h1>
    <p>Multi-Clinic · Upload files from any clinic · Cross-clinic fraud detection · Export findings</p>
  </div>
</div>
""", unsafe_allow_html=True)

if not st.session_state.df.empty:
    badges = "".join(
        f'<span class="clinic-badge">🏥 {c}</span>'
        for c in sorted(st.session_state.df["Clinic"].unique())
    )
    st.markdown(f"<div style='margin-bottom:.8rem'>{badges}</div>",
                unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📁 Upload Voucher Files")
    st.caption("Any clinic · Any month · Multiple files at once")

    uploaded = st.file_uploader(
        "Select Excel files",
        type=["xlsx", "xls"],
        accept_multiple_files=True,
        label_visibility="collapsed",
    )

    col1, col2 = st.columns(2)
    load_btn   = col1.button("🔄 Load", type="primary", use_container_width=True)
    clear_btn  = col2.button("🗑 Clear", use_container_width=True)

    if clear_btn:
        st.session_state.df           = pd.DataFrame()
        st.session_state.files_loaded = []
        st.cache_data.clear()
        st.rerun()

    if load_btn and uploaded:
        frames, errors = [], []
        prog = st.progress(0, text="Parsing…")
        for i, f in enumerate(uploaded):
            try:
                frames.append(parse_file(f.read(), f.name))
            except Exception as e:
                errors.append(f"{f.name}: {e}")
            prog.progress((i + 1) / len(uploaded), text=f"✔ {f.name}")

        if frames:
            combined = pd.concat(frames, ignore_index=True)
            combined.drop_duplicates(
                subset=["Voucher_ID", "Affiliation_No", "Date", "Clinic"],
                keep="first", inplace=True,
            )
            with st.spinner("Normalising affiliations…"):
                combined = build_canonical_affiliations(combined)
            with st.spinner("Detecting duplicates…"):
                combined = detect_duplicates(combined)
            st.session_state.df           = combined
            st.session_state.files_loaded = [f.name for f in uploaded]
            prog.empty()
            flagged = int((combined["Dup_Type"] != "").sum())
            st.success(f"✅ {len(combined):,} records · {flagged:,} flagged")

        for e in errors:
            st.warning(f"⚠ {e}")

    if st.session_state.files_loaded:
        st.markdown("---")
        st.markdown("**Loaded files:**")
        for fname in st.session_state.files_loaded:
            st.markdown(f"📄 `{fname}`")

    st.markdown("---")
    st.markdown("**Duplicate colour key:**")
    st.markdown("""
    <span class="chip chip-exact">⚠ Exact Duplicate</span><br>
    <span class="chip chip-voucher">⚠ Duplicate Voucher</span><br>
    <span class="chip chip-clinic">⚠ Cross-Clinic</span><br>
    <span class="chip chip-cross">ℹ Cross-Month</span><br>
    <span class="chip chip-month">ℹ Same ID / Month</span>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  EMPTY STATE
# ─────────────────────────────────────────────────────────────────────────────
df = st.session_state.df

if df.empty:
    st.markdown("""
    <div class="upload-hint">
      <h3>👈 Upload voucher files using the sidebar</h3>
      <p>Supports files from <strong>any clinic</strong>, mixed together.<br>
      Clinic name, district &amp; TIN are auto-detected from each file header.<br>
      <strong>Cross-clinic fraud signals</strong> are detected automatically when
      the same patient appears at multiple facilities.</p>
      <p style="margin-top:.8rem;font-size:.8rem;color:#58A6FF">
        📖 See the <strong>Guide</strong> tab after loading for a full explanation of all filters and duplicate types.
      </p>
    </div>
    """, unsafe_allow_html=True)
    # Still show the Guide tab even with no data loaded
    with t_guide:
        _render_guide()
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
#  GLOBAL CLINIC FILTER  (sidebar, affects all tabs)
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("---")
    st.markdown("**🔍 Filter by Clinic**")
    all_clinics = sorted(df["Clinic"].unique().tolist())
    sel_clinics = st.multiselect(
        "Active clinics", all_clinics, default=all_clinics,
        key="global_clinic_filter",
    )

df_view     = df[df["Clinic"].isin(sel_clinics)] if sel_clinics else df.copy()
flagged_df  = df_view[df_view["Dup_Type"] != ""]

# ─────────────────────────────────────────────────────────────────────────────
#  TABS
# ─────────────────────────────────────────────────────────────────────────────
t_dash, t_dups, t_expl, t_export, t_guide = st.tabs([
    "📊 Dashboard",
    "🚨 Duplicates",
    "🔎 Data Explorer",
    "📤 Export",
    "📖 Guide",
])

# ════════════════════════════════════════════════════════════════════════════
#  DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
with t_dash:

    # KPI row
    k1,k2,k3,k4,k5,k6,k7 = st.columns(7)
    k1.metric("🏥 Clinics",       df_view["Clinic"].nunique())
    k2.metric("📋 Total Claims",  f"{len(df_view):,}")
    k3.metric("💰 Total (100%)",  fmt(df_view["Total_100"].sum()))
    k4.metric("💵 Total (85%)",   fmt(df_view["Total_85"].sum()))
    k5.metric("🚨 Flagged",       f"{len(flagged_df):,}",
              delta=f"{len(flagged_df)/max(len(df_view),1)*100:.1f}%",
              delta_color="inverse")
    k6.metric("📅 Months",        df_view["Month_Label"].nunique())
    k7.metric("🏢 Affiliations",  df_view["Affiliation_Group"].nunique())

    st.markdown("---")

    # Row 1: Claims per month (grouped by clinic) + Service doughnut
    col_a, col_b = st.columns([3, 2])

    with col_a:
        st.markdown('<div class="section-title">Claims per Month — by Clinic</div>',
                    unsafe_allow_html=True)
        m_grp = (df_view.groupby(["Clinic", "Month_Label"])
                        .agg(Claims=("Voucher_ID","count"), Amount=("Total_100","sum"))
                        .reset_index())
        fig_m = px.bar(
            m_grp, x="Month_Label", y="Claims", color="Clinic",
            barmode="group", color_discrete_sequence=CHART_COLORS,
            labels={"Month_Label":"Month"},
            hover_data={"Amount": ":,.0f"},
        )
        fig_m.update_layout(
            margin=dict(l=0,r=0,t=10,b=0), height=300,
            legend=dict(orientation="h", y=1.12, font=dict(size=10, color="#C9D1D9")),
            plot_bgcolor="#161B22", paper_bgcolor="#161B22",
            font=dict(color="#C9D1D9"),
            xaxis=dict(gridcolor="#30363D", color="#8B949E"),
            yaxis=dict(gridcolor="#30363D", color="#8B949E"),
        )
        st.plotly_chart(fig_m, use_container_width=True)

    with col_b:
        st.markdown('<div class="section-title">Cost Breakdown by Service</div>',
                    unsafe_allow_html=True)
        svc_tot = {s: float(df_view[s].sum()) for s in SERVICE_COLS
                   if df_view[s].sum() > 0}
        fig_s = go.Figure(go.Pie(
            labels=list(svc_tot.keys()), values=list(svc_tot.values()),
            hole=.44, marker=dict(colors=CHART_COLORS), textinfo="percent",
            textfont=dict(color="#E6EDF3"),
            hovertemplate="%{label}: RWF %{value:,.0f}<extra></extra>",
        ))
        fig_s.update_layout(
            legend=dict(font=dict(size=10, color="#C9D1D9")),
            margin=dict(l=0,r=0,t=10,b=0), height=300,
            paper_bgcolor="#161B22", font=dict(color="#C9D1D9"),
        )
        st.plotly_chart(fig_s, use_container_width=True)

    # Row 2: Amount per clinic + Dup breakdown
    col_c, col_d = st.columns([3, 2])

    with col_c:
        st.markdown('<div class="section-title">Total Claim Amount per Clinic</div>',
                    unsafe_allow_html=True)
        cgrp = (df_view.groupby("Clinic")
                       .agg(Amount_100=("Total_100","sum"),
                            Amount_85=("Total_85","sum"),
                            Claims=("Voucher_ID","count"),
                            Flagged=("Dup_Type", lambda x: (x!="").sum()))
                       .reset_index()
                       .sort_values("Amount_100", ascending=False))
        fig_c = go.Figure()
        fig_c.add_trace(go.Bar(name="Total (100%)", x=cgrp["Clinic"],
                               y=cgrp["Amount_100"], marker_color="#1F6FEB"))
        fig_c.add_trace(go.Bar(name="Total (85%)",  x=cgrp["Clinic"],
                               y=cgrp["Amount_85"],  marker_color="#3FB950"))
        fig_c.update_layout(
            barmode="group",
            margin=dict(l=0,r=0,t=10,b=0), height=320,
            legend=dict(orientation="h",y=1.1,font=dict(size=10, color="#C9D1D9")),
            plot_bgcolor="#161B22", paper_bgcolor="#161B22",
            font=dict(color="#C9D1D9"),
            xaxis=dict(gridcolor="#30363D", color="#8B949E"),
            yaxis=dict(title="Amount (RWF)", gridcolor="#30363D", color="#8B949E"),
        )
        st.plotly_chart(fig_c, use_container_width=True)

    with col_d:
        st.markdown('<div class="section-title">Duplicate Type Breakdown</div>',
                    unsafe_allow_html=True)
        if not flagged_df.empty:
            dbd = flagged_df["Dup_Type"].value_counts().reset_index()
            dbd.columns = ["Type","Count"]
            fig_d = go.Figure(go.Pie(
                labels=dbd["Type"], values=dbd["Count"], hole=.44,
                marker=dict(colors=[DUP_COLORS.get(t,"#AAA") for t in dbd["Type"]]),
                textinfo="value+percent",
                textfont=dict(color="#E6EDF3"),
                hovertemplate="%{label}: %{value}<extra></extra>",
            ))
            fig_d.update_layout(
                legend=dict(font=dict(size=9, color="#C9D1D9")),
                margin=dict(l=0,r=0,t=10,b=0), height=320,
                paper_bgcolor="#161B22", font=dict(color="#C9D1D9"),
            )
            st.plotly_chart(fig_d, use_container_width=True)
        else:
            st.info("No duplicates in selected data.")

    # Row 3: Top affiliations + Flagged heatmap
    st.markdown("---")
    col_e, col_f = st.columns([1, 1])

    with col_e:
        st.markdown('<div class="section-title">Top 10 Affiliations by Claim Amount</div>',
                    unsafe_allow_html=True)
        top10 = (df_view.groupby("Affiliation_Group")["Total_100"]
                        .sum().nlargest(10).reset_index())
        top10.columns = ["Affiliation","Amount"]
        fig_t = px.bar(
            top10, x="Amount", y="Affiliation", orientation="h",
            color="Amount", color_continuous_scale=["#1C2A3A","#1F6FEB"],
            text=top10["Amount"].apply(lambda v: f"RWF {int(v):,}"),
        )
        fig_t.update_traces(textposition="outside", textfont=dict(color="#C9D1D9"))
        fig_t.update_layout(
            coloraxis_showscale=False,
            margin=dict(l=0,r=0,t=10,b=0), height=360,
            plot_bgcolor="#161B22", paper_bgcolor="#161B22",
            font=dict(color="#C9D1D9"),
            xaxis=dict(gridcolor="#30363D", color="#8B949E"),
            yaxis=dict(autorange="reversed", color="#8B949E"),
        )
        st.plotly_chart(fig_t, use_container_width=True)

    with col_f:
        st.markdown('<div class="section-title">Flagged Entries — Clinic × Month Heatmap</div>',
                    unsafe_allow_html=True)
        if not flagged_df.empty:
            hd = (flagged_df.groupby(["Clinic","Month_Label"])
                            .size().reset_index(name="Count"))
            pivot = hd.pivot_table(index="Clinic", columns="Month_Label",
                                   values="Count", fill_value=0)
            fig_h = px.imshow(
                pivot, color_continuous_scale="Reds",
                aspect="auto", text_auto=True, labels={"color":"Flagged"},
            )
            fig_h.update_layout(
                margin=dict(l=0,r=0,t=10,b=0), height=360,
                paper_bgcolor="#161B22", font=dict(color="#C9D1D9"),
                xaxis=dict(color="#8B949E"), yaxis=dict(color="#8B949E"),
            )
            st.plotly_chart(fig_h, use_container_width=True)
        else:
            st.info("No flagged entries to display.")

    # Clinic summary table
    st.markdown("---")
    st.markdown('<div class="section-title">Clinic Summary</div>',
                unsafe_allow_html=True)
    cs = (df_view.groupby(["Clinic","District","TIN"]).agg(
            Files        = ("Source_File",     "nunique"),
            Months       = ("Month_Label",     "nunique"),
            Claims       = ("Voucher_ID",      "count"),
            Total_100    = ("Total_100",        "sum"),
            Total_85     = ("Total_85",         "sum"),
            Flagged      = ("Dup_Type",         lambda x: (x!="").sum()),
            Affiliations = ("Affiliation_Group","nunique"),
          ).reset_index())
    cs["Total_100"] = cs["Total_100"].apply(lambda v: f"RWF {int(v):,}")
    cs["Total_85"]  = cs["Total_85"].apply(lambda v: f"RWF {int(v):,}")
    cs["Flag %"]    = cs.apply(
        lambda r: f"{r['Flagged']/max(r['Claims'],1)*100:.1f}%", axis=1
    )
    st.dataframe(cs, use_container_width=True, hide_index=True)


# ════════════════════════════════════════════════════════════════════════════
#  DUPLICATES
# ════════════════════════════════════════════════════════════════════════════
with t_dups:
    st.markdown('<div class="section-title">🚨 Flagged Duplicate Entries</div>',
                unsafe_allow_html=True)

    dc1,dc2,dc3,dc4 = st.columns([2,2,2,2])
    d_type   = dc1.selectbox("Type",   ["All"] + DUP_ORDER,                             key="dt")
    d_month  = dc2.selectbox("Month",  ["All"] + sorted(df_view["Month_Label"].unique()),key="dm")
    d_clinic = dc3.selectbox("Clinic", ["All"] + sorted(df_view["Clinic"].unique()),     key="dc")
    d_name   = dc4.text_input("Name search", key="dn")

    dv = flagged_df.copy()
    if d_type   != "All": dv = dv[dv["Dup_Type"]    == d_type]
    if d_month  != "All": dv = dv[dv["Month_Label"] == d_month]
    if d_clinic != "All": dv = dv[dv["Clinic"]      == d_clinic]
    if d_name.strip():
        dv = dv[dv["Beneficiary_Name"].str.contains(d_name.strip().upper(), na=False)]

    st.caption(
        f"**{len(dv):,}** flagged rows"
        + (f" (from {len(flagged_df):,} total)" if len(dv) < len(flagged_df) else "")
    )

    dup_show_cols = [
        "Date","Voucher_ID","Affiliation_No","Beneficiary_Name",
        "Clinic","District","Affiliation_Group",
        "Total_100","Total_85","Dup_Type","Month_Label","Source_File",
    ]
    dv_show = dv[dup_show_cols].copy()
    dv_show.rename(columns={
        "Affiliation_Group":"Affiliation","Month_Label":"Month",
        "Source_File":"File","Total_100":"Total(100%)","Total_85":"Total(85%)",
    }, inplace=True)
    st.dataframe(
        dv_show.style.apply(highlight_dups, axis=1),
        use_container_width=True, height=520,
    )

    # Cross-clinic deep-dive
    cc = flagged_df[flagged_df["Dup_Type"] == "Cross-Clinic Duplicate"]
    if not cc.empty:
        with st.expander(f"🔴 Cross-Clinic Deep Dive — {len(cc):,} entries across clinics"):
            st.caption(
                "Same patient (Affiliation No + Name) billed at multiple clinics. "
                "Sorted by patient so all entries per person are grouped."
            )
            cc_show = (cc[["Affiliation_No","Beneficiary_Name","Date",
                            "Clinic","Voucher_ID","Total_100","Month_Label"]]
                         .sort_values(["Affiliation_No","Beneficiary_Name","Date"])
                         .copy())
            cc_show["Total_100"] = cc_show["Total_100"].apply(lambda v: f"RWF {int(v):,}")
            cc_show.rename(columns={"Month_Label":"Month","Total_100":"Total(100%)"}, inplace=True)
            st.dataframe(cc_show, use_container_width=True, hide_index=True)


# ════════════════════════════════════════════════════════════════════════════
#  DATA EXPLORER
# ════════════════════════════════════════════════════════════════════════════
with t_expl:
    st.markdown('<div class="section-title">🔎 Explore All Claims</div>',
                unsafe_allow_html=True)

    ef1,ef2,ef3,ef4,ef5,ef6 = st.columns([2,2,2,2,2,2])
    e_clinic = ef1.selectbox("Clinic",
        ["All"] + sorted(df_view["Clinic"].unique()), key="ec")
    e_month  = ef2.selectbox("Month",
        ["All"] + sorted(df_view["Month_Label"].unique()), key="em")
    e_svc    = ef3.selectbox("Service Used",
        ["All"] + SERVICE_COLS, key="es")
    e_aff    = ef4.selectbox("Affiliation",
        ["All"] + sorted(df_view["Affiliation_Group"].dropna().unique()), key="ea")
    e_flag   = ef5.selectbox("Flag Status",
        ["All","Flagged Only","Clean Only"] + DUP_ORDER, key="ef")
    e_name   = ef6.text_input("Name search", key="en")

    expl = df_view.copy()
    if e_clinic != "All": expl = expl[expl["Clinic"]            == e_clinic]
    if e_month  != "All": expl = expl[expl["Month_Label"]       == e_month]
    if e_svc    != "All": expl = expl[expl[e_svc]               >  0]
    if e_aff    != "All": expl = expl[expl["Affiliation_Group"] == e_aff]
    if e_flag == "Flagged Only": expl = expl[expl["Dup_Type"] != ""]
    elif e_flag == "Clean Only": expl = expl[expl["Dup_Type"] == ""]
    elif e_flag in DUP_ORDER:    expl = expl[expl["Dup_Type"] == e_flag]
    if e_name.strip():
        expl = expl[expl["Beneficiary_Name"].str.contains(
            e_name.strip().upper(), na=False)]

    st.caption(
        f"**{len(expl):,}** records"
        + (f" (filtered from {len(df_view):,})" if len(expl) < len(df_view) else "")
    )

    show_cols = [
        "Date","Voucher_ID","Affiliation_No","Beneficiary_Name",
        "Clinic","District","Affiliation_Group",
        "Consultation","Laboratory","Imaging","Hospitalization",
        "Procedures","Consumables","Medicines",
        "Total_100","Total_85","Dup_Type","Month_Label",
    ]
    expl_show = expl[show_cols].copy()
    expl_show.rename(columns={
        "Affiliation_Group":"Affiliation","Month_Label":"Month",
        "Total_100":"Total(100%)","Total_85":"Total(85%)",
    }, inplace=True)
    st.dataframe(
        expl_show.style.apply(highlight_dups, axis=1),
        use_container_width=True, height=540,
    )


# ════════════════════════════════════════════════════════════════════════════
#  EXPORT
# ════════════════════════════════════════════════════════════════════════════
with t_export:
    st.markdown('<div class="section-title">📤 Export Report to Excel</div>',
                unsafe_allow_html=True)

    col_l, col_r = st.columns([3, 1])
    n_flagged_exp = int((df_view["Dup_Type"] != "").sum())
    n_cc_exp      = int((df_view["Dup_Type"] == "Cross-Clinic Duplicate").sum())

    with col_l:
        st.markdown(f"""
The exported workbook will contain **5 colour-coded sheets**:

| Sheet | Contents |
|-------|----------|
| 📋 **Full Data** | All **{len(df_view):,}** records — highlighted by duplicate type |
| 🚨 **Duplicates** | **{n_flagged_exp:,}** flagged entries with type and reason |
| 🏢 **By Affiliation** | Claim totals per affiliation group, per clinic |
| 📅 **By Month** | Monthly totals per clinic |
| 🔴 **Cross-Clinic Flags** | **{n_cc_exp:,}** patients seen at multiple facilities |

**Row colour coding:**
        """)
        for label, color in DUP_COLORS.items():
            st.markdown(
                f'<span style="background:{color};color:#ffffff;font-weight:600;'
                f'padding:.3rem 1rem;border-radius:6px;font-size:.82rem;'
                f'margin:.2rem;display:inline-block">'
                f'{label}</span>',
                unsafe_allow_html=True,
            )

    with col_r:
        st.markdown("<br><br>", unsafe_allow_html=True)
        with st.spinner("Building workbook…"):
            xlsx_bytes = build_export(df_view)
        st.download_button(
            label="📥 Download Excel Report",
            data=xlsx_bytes,
            file_name="Rwanda_Voucher_Fraud_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )
        st.caption(
            f"{len(df_view):,} rows · "
            f"{df_view['Clinic'].nunique()} clinic(s) · "
            f"{n_flagged_exp:,} flagged"
        )


# ════════════════════════════════════════════════════════════════════════════
#  GUIDE
# ════════════════════════════════════════════════════════════════════════════
with t_guide:
    _render_guide()